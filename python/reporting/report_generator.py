"""
python/reporting/report_generator.py
--------------------------------------
Stage 3b: Automated Excel report generation from the reporting layer.

Generates a multi-tab Excel workbook:
    - Executive Summary    (revenue, profit, growth KPIs)
    - Monthly Trends       (month-by-month with MoM / YoY)
    - Regional Performance (by region and year)
    - Product Analysis     (by category and product)
    - Anomaly Report       (flagged transactions)

Output: data/processed/report_summary.xlsx

Usage:
    python -m python.reporting.report_generator
"""

import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

from python.utils.db_connector import DBConnector
from python.utils.config_loader import load_config

logger = logging.getLogger(__name__)

# ── Style Constants ──────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # Dark navy
SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")   # Medium blue
ALT_ROW_FILL = PatternFill("solid", fgColor="EBF3FB")   # Light blue
GREEN_FILL   = PatternFill("solid", fgColor="E2EFDA")
RED_FILL     = PatternFill("solid", fgColor="FFDEDE")

HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Calibri", size=10)
TITLE_FONT   = Font(name="Calibri", bold=True, size=14, color="1F3864")

THIN_BORDER  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

CENTER = Alignment(horizontal="center", vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")


class ReportGenerator:
    """Generates automated Excel reports from the BI reporting tables."""

    def __init__(self, config: dict = None):
        self.config   = config or load_config()
        self.db       = DBConnector(self.config["database"])
        self.out_path = self.config["reporting"]["output_path"]
        os.makedirs(self.out_path, exist_ok=True)

    # ── Public Interface ─────────────────────────────────────────────────────

    def run(self) -> str:
        """
        Generate the full Excel report.

        Returns:
            Path to the generated file.
        """
        logger.info("=" * 60)
        logger.info("STAGE 3b: Report Generation")
        logger.info("=" * 60)

        data = self._fetch_all_data()
        filepath = self._build_excel(data)
        logger.info(f"✅ Report saved: {filepath}")
        return filepath

    # ── Data Fetching ─────────────────────────────────────────────────────────

    def _fetch_all_data(self) -> dict:
        """Load all reporting tables into DataFrames."""
        logger.info("Fetching data from reporting layer...")
        self.db.connect()

        data = {
            "monthly":   self.db.query_df("SELECT * FROM vw_revenue_trends ORDER BY txn_year, txn_month"),
            "regional":  self.db.query_df("SELECT * FROM vw_regional_performance ORDER BY txn_year, total_revenue DESC"),
            "product":   self.db.query_df("SELECT * FROM rpt_product_summary ORDER BY txn_year, total_revenue DESC"),
            "anomalies": self.db.query_df("SELECT * FROM rpt_anomalies ORDER BY revenue_zscore DESC NULLS LAST"),
            "core":      self.db.query_df("SELECT txn_year, SUM(revenue) AS rev, SUM(gross_profit) AS profit, COUNT(*) AS txns FROM core_sales GROUP BY txn_year"),
        }

        self.db.disconnect()
        logger.info(f"  Monthly rows:   {len(data['monthly']):,}")
        logger.info(f"  Regional rows:  {len(data['regional']):,}")
        logger.info(f"  Product rows:   {len(data['product']):,}")
        logger.info(f"  Anomaly rows:   {len(data['anomalies']):,}")
        return data

    # ── Excel Builder ─────────────────────────────────────────────────────────

    def _build_excel(self, data: dict) -> str:
        """Write all report tabs to Excel."""
        filename = f"report_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.out_path, filename)

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            self._write_executive_summary(writer, data)
            self._write_monthly_trends(writer, data["monthly"])
            self._write_regional(writer, data["regional"])
            self._write_products(writer, data["product"])
            self._write_anomalies(writer, data["anomalies"])

        # Post-process: apply styling
        self._apply_styles(filepath)
        return filepath

    def _write_executive_summary(self, writer, data: dict) -> None:
        """Tab 1: Executive KPI summary."""
        core = data["core"]
        rows = []
        for _, r in core.iterrows():
            rows.append({
                "Year":              int(r["txn_year"]),
                "Total Revenue ($)": round(r["rev"], 2),
                "Gross Profit ($)":  round(r["profit"], 2),
                "Profit Margin (%)": round(r["profit"] / r["rev"] * 100, 1) if r["rev"] else 0,
                "Transactions":      int(r["txns"]),
            })

        # Add YoY growth column
        df = pd.DataFrame(rows)
        df["YoY Revenue Growth (%)"] = df["Total Revenue ($)"].pct_change() * 100
        df["YoY Revenue Growth (%)"] = df["YoY Revenue Growth (%)"].round(1)

        df.to_excel(writer, sheet_name="Executive Summary", index=False, startrow=2)
        logger.info("  Tab: Executive Summary")

    def _write_monthly_trends(self, writer, df: pd.DataFrame) -> None:
        """Tab 2: Monthly revenue trends."""
        cols = [
            "month_label", "txn_year", "txn_month",
            "total_revenue", "total_gross_profit", "profit_margin_pct",
            "transaction_count", "mom_growth_pct", "yoy_growth_pct", "ytd_revenue"
        ]
        df[cols].to_excel(writer, sheet_name="Monthly Trends", index=False, startrow=1)
        logger.info("  Tab: Monthly Trends")

    def _write_regional(self, writer, df: pd.DataFrame) -> None:
        """Tab 3: Regional performance."""
        cols = [
            "region", "txn_year", "total_revenue", "total_gross_profit",
            "profit_margin_pct", "transaction_count",
            "yoy_revenue_growth_pct", "region_revenue_share_pct",
            "avg_transaction_value", "top_sales_rep"
        ]
        available = [c for c in cols if c in df.columns]
        df[available].to_excel(writer, sheet_name="Regional Performance", index=False, startrow=1)
        logger.info("  Tab: Regional Performance")

    def _write_products(self, writer, df: pd.DataFrame) -> None:
        """Tab 4: Product & category breakdown."""
        df.to_excel(writer, sheet_name="Product Analysis", index=False, startrow=1)
        logger.info("  Tab: Product Analysis")

    def _write_anomalies(self, writer, df: pd.DataFrame) -> None:
        """Tab 5: Anomaly detection report."""
        if df.empty:
            pd.DataFrame({"message": ["No anomalies detected."]}).to_excel(
                writer, sheet_name="Anomaly Report", index=False
            )
        else:
            df.to_excel(writer, sheet_name="Anomaly Report", index=False, startrow=1)
        logger.info("  Tab: Anomaly Report")

    def _apply_styles(self, filepath: str) -> None:
        """Apply professional formatting to the generated workbook."""
        wb = load_workbook(filepath)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Style header row (row 1 or 2 depending on tab)
            header_row = 2 if sheet_name == "Executive Summary" else 2
            for cell in ws[header_row]:
                if cell.value:
                    cell.fill   = HEADER_FILL
                    cell.font   = HEADER_FONT
                    cell.alignment = CENTER
                    cell.border = THIN_BORDER

            # Auto-fit columns
            for col in ws.columns:
                max_len = max(
                    (len(str(cell.value)) for cell in col if cell.value),
                    default=10
                )
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

            # Freeze header row
            ws.freeze_panes = f"A{header_row + 1}"

        wb.save(filepath)
        logger.info("  Styles applied to workbook.")


if __name__ == "__main__":
    from python.utils.logger import setup_logger
    setup_logger()
    gen = ReportGenerator()
    gen.run()
